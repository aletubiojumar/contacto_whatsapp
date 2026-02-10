#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import os
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def load_dotenv(explicit_path: str | None = None) -> Path | None:
    candidates: list[Path] = []

    if explicit_path:
        candidates.append(Path(explicit_path).expanduser())

    candidates.append(Path.cwd() / ".env")
    script_dir = Path(__file__).resolve().parent
    candidates.append(script_dir / ".env")
    for i in range(1, 5):
        candidates.append(script_dir.parents[i - 1] / ".env")

    env_path = next((p for p in candidates if p.exists()), None)
    if not env_path:
        return None

    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())

    return env_path


HEADERS = [
    "Encargo",
    "Fecha Sin.",
    "Causa",
    "Aseguradora",
]

QUERY = r"""
WITH
-- 1) Siniestros que "aparecen" hoy en PL por tener algún encargo con fecha de HOY
siniestros_en_rango AS (
  SELECT
    e.id_siniestro,
    MAX(e.encargo_fh) AS encargo_fh_rango,
    MAX(CASE WHEN e.id_causa != 15 THEN e.id_causa ELSE NULL END) AS id_causa  -- Excluir "Asistencia a Juicio"
  FROM softline_encargos e
  WHERE e.id_despacho = 2
    AND e.encargo_fh >= CURDATE()
    AND e.encargo_fh <  (CURDATE() + INTERVAL 1 DAY)
  GROUP BY e.id_siniestro
),

-- 2) Histórico de encargos SOLO de esos siniestros (para calcular nº encargos y TTR_1/TTR_2)
enc_rank AS (
  SELECT
    e.id_siniestro,
    e.id_encargo,
    ROW_NUMBER() OVER (
      PARTITION BY e.id_siniestro
      ORDER BY COALESCE(e.encargo_num, 9999), e.encargo_fh, e.id_encargo
    ) AS rn,
    CASE
      WHEN (COALESCE(e.encargo_estado_ttr,0) > 0 OR e.encargo_tipo = 12)
      THEN 1 ELSE 0
    END AS is_ttr
  FROM softline_encargos e
  JOIN siniestros_en_rango sr ON sr.id_siniestro = e.id_siniestro
  WHERE e.id_despacho = 2
),

-- 3) Agregados por siniestro (histórico)
enc_info AS (
  SELECT
    id_siniestro,
    COUNT(*) AS n_encargos,
    MAX(CASE WHEN rn=1 THEN is_ttr ELSE 0 END) AS ttr_1,
    MAX(CASE WHEN rn=2 THEN is_ttr ELSE 0 END) AS ttr_2
  FROM enc_rank
  GROUP BY id_siniestro
)

SELECT
  c.codigo AS `ENCARGO`,
  DATE_FORMAT(s.siniestro_fh, '%d/%m/%Y') AS `FECHA SIN.`,
  COALESCE(mc.descripcion, '') AS `CAUSA`,
  CASE 
    WHEN s.id_cia = 42 THEN 'Allianz'
    WHEN s.id_cia = 399 THEN 'AllianzBBVA'
    ELSE 'Desconocida'
  END AS `ASEGURADORA`
FROM softline_siniestros s
JOIN siniestros_en_rango sr ON sr.id_siniestro = s.id_siniestro
JOIN softline_siniestros_codigos c
  ON c.id_codigo = s.id_cod_siniestro
 AND c.tipo_codigo = 'siniestro'
LEFT JOIN softline_maestra_causas mc
  ON mc.id_cod = sr.id_causa
JOIN enc_info ei ON ei.id_siniestro = s.id_siniestro
WHERE s.id_despacho = 2
  AND s.id_cia IN (42,399)
  AND CHAR_LENGTH(c.codigo) >= 9

  -- sin contacto (criterio actual)
  AND NOT EXISTS (
    SELECT 1
    FROM softline_encargos e
    JOIN softline_encargos_contactos ec
      ON ec.contactos_id_encargo = e.id_encargo
    WHERE e.id_siniestro = s.id_siniestro
      AND ec.contactos_fechahora IS NOT NULL
  )

  -- regla diagrama aplicada al histórico
  AND (
    (ei.n_encargos = 1 AND ei.ttr_1 = 0)
    OR
    (ei.n_encargos = 2 AND ei.ttr_1 = 1 AND ei.ttr_2 = 0)
  )
ORDER BY sr.encargo_fh_rango DESC, s.id_siniestro DESC
"""


@dataclass
class DBConfig:
    host: str
    port: int
    user: str
    password: str
    dbname: str
    ssl_ca: str | None
    ssl_cert: str | None
    ssl_key: str | None


def read_db_config() -> DBConfig:
    def req(name: str) -> str:
        v = os.getenv(name)
        if not v:
            raise SystemExit(f"Falta variable de entorno obligatoria: {name}")
        return v

    return DBConfig(
        host=req("DB_HOST"),
        port=int(os.getenv("DB_PORT", "3306")),
        user=req("DB_USER"),
        password=req("DB_PASS"),
        dbname=req("DB_NAME"),
        ssl_ca=os.getenv("DB_SSL_CA"),
        ssl_cert=os.getenv("DB_SSL_CERT"),
        ssl_key=os.getenv("DB_SSL_KEY"),
    )


def mariadb_cmd(cfg: DBConfig) -> List[str]:
    cmd = [
        "mariadb",
        f"--host={cfg.host}",
        f"--port={cfg.port}",
        f"--user={cfg.user}",
        f"--password={cfg.password}",
        "--database",
        cfg.dbname,
        "--batch",
        "--raw",
        "--skip-column-names",
    ]

    if cfg.ssl_ca or cfg.ssl_cert or cfg.ssl_key:
        cmd.append("--ssl")
        if cfg.ssl_ca:
            cmd.append(f"--ssl-ca={cfg.ssl_ca}")
        if cfg.ssl_cert:
            cmd.append(f"--ssl-cert={cfg.ssl_cert}")
        if cfg.ssl_key:
            cmd.append(f"--ssl-key={cfg.ssl_key}")

    return cmd


def run_query(cfg: DBConfig) -> List[Tuple[str, str, str, str]]:
    cmd = mariadb_cmd(cfg) + ["--execute", QUERY]
    try:
        output = subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True)
    except FileNotFoundError:
        raise SystemExit("No se encuentra el ejecutable 'mariadb'")
    except subprocess.CalledProcessError as e:
        raise SystemExit(f"Error ejecutando la query:\n{e.output}")

    rows: List[Tuple[str, str, str, str]] = []
    for line in output.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split("\t")
        if len(parts) >= 4:
            encargo = parts[0].strip()
            fecha = parts[1].strip()
            causa = parts[2].strip()
            aseguradora = parts[3].strip()
            rows.append((encargo, fecha, causa, aseguradora))
    return rows


def write_excel(rows: List[Tuple[str, str, str, str]], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worksheet"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for i, (encargo, fecha, causa, aseguradora) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=encargo)
        c = ws.cell(row=i, column=1)
        c.number_format = "@"
        
        ws.cell(row=i, column=2, value=fecha)
        ws.cell(row=i, column=3, value=causa)
        ws.cell(row=i, column=4, value=aseguradora)

    for col in range(1, len(HEADERS) + 1):
        max_len = 0
        for r in range(1, min(ws.max_row, 2000) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True, help="Ruta del Excel de salida")
    parser.add_argument("--env", default=None, help="Ruta opcional a un archivo .env")
    args = parser.parse_args()

    loaded = load_dotenv(args.env)
    if loaded:
        print(f".env cargado: {loaded}")

    cfg = read_db_config()
    rows = run_query(cfg)
    write_excel(rows, Path(args.out).expanduser().resolve())

    print(f"OK: {len(rows)} filas exportadas → {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())