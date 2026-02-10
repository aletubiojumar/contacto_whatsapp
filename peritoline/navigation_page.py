"""
Extrae teléfonos desde ePAC para siniestros Allianz.
Modos:
- --refresh: borra excels existentes (raw_allianz) y descarga de nuevo desde Peritoline
- --excel: usa un Excel existente (sin descargar)
- --headless: ejecuta con navegador oculto (sin UI). Si no, con navegador visible.
Flujo:
1) Obtener Excel (descargar o usar existente)
2) Filtrar Excel (genera *_filtrado.xlsx)
3) Obtener credenciales ePAC desde la BASE DE DATOS (ya no usa Playwright)
4) Login ePAC -> ir a Peritaciones Diversos
5) Por cada siniestro:
   - asegurar pantalla de búsqueda (#claimNumber visible)
   - buscar, seleccionar
   - menú lateral: HSC y FP -> Ficha peritación
   - extraer teléfono (obs -> descripción -> telef-1 -> telef-2)
   - volver a Peritaciones Diversos
6) Guardar salida TXT
"""
from __future__ import annotations
import argparse
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from playwright.sync_api import Page

from browser import launch_browser
from config import load_config, AppConfig
from peritoline.downloader import download_report_in_session
from peritoline.login_page import PeritolineLoginPage
from peritoline.filter_excel import filter_excel
from epac.pages.num_siniestro_page import NumeroSiniestroPage
from epac.pages.epac_ficha_peritacion_page import EpacFichaPeritacionPage
from utils.human import human_delay
from utils.logging_utils import get_logger, setup_logging

# Importar mysql.connector para obtener credenciales de BD
try:
    import mysql.connector  # type: ignore
except ImportError:
    mysql = None
    mysql_import_error = ImportError("mysql.connector no disponible")
else:
    mysql_import_error = None

RAW_DIR = Path("data/peritoline/raw_allianz")

# -----------------------------------------------------------------------------
# Utilidades Excel
# -----------------------------------------------------------------------------
def borrar_excels_raw_allianz() -> int:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    n = 0
    for p in RAW_DIR.glob("*.xlsx"):
        try:
            p.unlink()
            n += 1
        except Exception:
            pass
    for p in RAW_DIR.glob("*.xls"):
        try:
            p.unlink()
            n += 1
        except Exception:
            pass
    return n

def pick_latest_excel(raw_dir: Path) -> Optional[Path]:
    files = [p for p in raw_dir.glob("*.xlsx") if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)

def normalizar_siniestro(s: str) -> str:
    return re.sub(r"\D", "", (s or "").strip())

def filtrar_siniestros_validos(lista: list[str], min_len: int = 9) -> list[str]:
    out = []
    seen = set()
    for s in lista:
        x = normalizar_siniestro(s)
        if len(x) < min_len:
            continue
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

# -----------------------------------------------------------------------------
# Credenciales ePAC desde BASE DE DATOS
# -----------------------------------------------------------------------------
def obtener_credenciales_epac(config: AppConfig) -> dict[str, str]:
    """
    Obtiene las credenciales de ePAC directamente desde la base de datos.
    
    Args:
        config: Configuración de la aplicación (para obtener datos de BD)
    
    Returns:
        dict con 'url', 'username', 'password'
    """
    logger = get_logger()
    logger.info(
        "Siniestro: sin_codigo | Tarea: obtener_credenciales_epac | "
        "Obteniendo credenciales de ePAC desde la base de datos"
    )
    
    # Verificar que mysql.connector esté disponible
    if mysql_import_error is not None:
        raise RuntimeError(
            "mysql.connector no disponible. Instala con: "
            "pip install mysql-connector-python"
        )
    
    # Conectar a la base de datos usando config
    try:
        cnx = mysql.connector.connect(
            host=config.db_host,
            port=config.db_port,
            user=config.db_user,
            password=config.db_password,
            database=config.db_name,
        )
    except Exception as e:
        logger.error(f"Error conectando a la base de datos: {e}")
        raise RuntimeError(f"No se pudo conectar a la base de datos: {e}")
    
    # Query para obtener credenciales de Allianz (id_cia = 42)
    query = """
    SELECT 
        url,
        user as usuario,
        pass as password
    FROM softline_aseguradoras_claves_web
    WHERE id_cia = %s
    LIMIT 1
    """
    
    try:
        cur = cnx.cursor(dictionary=True)
        cur.execute(query, (42,))  # 42 es el id_cia de Allianz
        result = cur.fetchone()
        cur.close()
        
        if not result:
            raise ValueError("No se encontraron credenciales para Allianz (id_cia=42)")
        
        # Verificar que tengamos los datos necesarios
        if not all([result.get('url'), result.get('usuario'), result.get('password')]):
            raise ValueError("Credenciales incompletas para Allianz")
        
        logger.info(f"✓ Credenciales obtenidas: Usuario={result['usuario']}")
        
        return {
            'url': result['url'],
            'username': result['usuario'],
            'password': result['password']
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo credenciales: {e}")
        raise
    finally:
        try:
            cnx.close()
        except Exception:
            pass

# -----------------------------------------------------------------------------
# Obtener Excel
# -----------------------------------------------------------------------------
def obtener_excel_allianz(config: AppConfig, refresh: bool, excel_path: Optional[str]) -> Path:
    logger = get_logger()
    if excel_path:
        p = Path(excel_path)
        if not p.exists():
            raise FileNotFoundError(f"No existe {p}")
        logger.info(f"Usando Excel manual: {p}")
        return p
    if refresh:
        n = borrar_excels_raw_allianz()
        logger.info(f"Borrados {n} excels antiguos en {RAW_DIR}")
    latest = pick_latest_excel(RAW_DIR)
    if latest:
        logger.info(f"Excel existente encontrado: {latest}")
        return latest
    logger.info("No hay Excel previo. Descargando desde Peritoline...")
    with launch_browser(config) as (context, page):
        login_page = PeritolineLoginPage(page, config)
        login_page.navigate()
        login_page.login()
        out_path = download_report_in_session(page, config)
    logger.info(f"Descarga completa: {out_path}")
    return out_path

def aplicar_filtro_excel(excel_in: Path) -> Path:
    logger = get_logger()
    out_path = excel_in.parent / (excel_in.stem + "_filtrado.xlsx")
    n_elim, n_ok = filter_excel(excel_in, out_path)
    logger.info(f"✓ Excel filtrado guardado en: {out_path}")
    logger.info(f"  - Filas eliminadas: {n_elim}")
    logger.info(f"  - Siniestros válidos: {n_ok}")
    return out_path

# -----------------------------------------------------------------------------
# Extracción de teléfonos
# -----------------------------------------------------------------------------
def extraer_telefonos_epac(
    config: AppConfig,
    siniestros: list[str],
    credenciales: dict[str, str],
) -> dict[str, str]:
    logger = get_logger()
    resultados: dict[str, str] = {}
    with launch_browser(config) as (context, epac_page):
        epac_page.goto(credenciales["url"], wait_until="domcontentloaded")
        num_sin_page = NumeroSiniestroPage(epac_page, config)
        num_sin_page.login(credenciales["username"], credenciales["password"])
        num_sin_page.ir_a_peritaciones_diversos()
        for idx, sin_code in enumerate(siniestros, start=1):
            logger.info(f"[{idx}/{len(siniestros)}] Procesando {sin_code}...")
            try:
                num_sin_page.asegurar_pantalla_busqueda()
                num_sin_page.buscar_siniestro(sin_code)
                num_sin_page.seleccionar_siniestro(sin_code)
                num_sin_page.ir_a_hsc_y_fp()
                ficha = EpacFichaPeritacionPage(epac_page, config)
                ficha.ir_a_ficha_peritacion()
                telefono = ficha.extraer_telefono()
                resultados[sin_code] = telefono
                logger.info(f"  → Teléfono: {telefono}")
                num_sin_page.volver_a_peritaciones_diversos()
            except Exception as e:
                logger.error(f"  ✗ Error procesando {sin_code}: {e}")
                resultados[sin_code] = "ERROR"
    return resultados

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--refresh", action="store_true", help="Borrar excels antiguos y descargar")
    parser.add_argument("--excel", help="Ruta a un Excel existente")
    parser.add_argument("--headless", action="store_true", help="Navegador oculto")
    args = parser.parse_args()
    
    setup_logging()
    logger = get_logger()
    config = load_config()
    config.headless = args.headless
    
    try:
        # 1) Obtener Excel
        excel_path = obtener_excel_allianz(config, args.refresh, args.excel)
        
        # 2) Filtrar Excel
        excel_filtrado = aplicar_filtro_excel(excel_path)
        
        # 3) Leer siniestros del Excel filtrado
        from openpyxl import load_workbook
        wb = load_workbook(excel_filtrado)
        ws = wb.active
        siniestros_raw = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3:
                siniestros_raw.append(str(row[2]))
        siniestros = filtrar_siniestros_validos(siniestros_raw)
        logger.info(f"  - Siniestros válidos: {len(siniestros)}")
        
        if not siniestros:
            logger.warning("No hay siniestros para procesar")
            return 0
        
        # 4) Obtener credenciales desde BD (YA NO USA PLAYWRIGHT)
        credenciales = obtener_credenciales_epac(config)
        
        # 5) Extraer teléfonos
        resultados = extraer_telefonos_epac(config, siniestros, credenciales)
        
        # 6) Guardar resultados
        out_txt = RAW_DIR / f"telefonos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(out_txt, "w", encoding="utf-8") as f:
            for sin_code, telefono in resultados.items():
                f.write(f"{sin_code}\t{telefono}\n")
        
        logger.info(f"✓ Resultados guardados en: {out_txt}")
        logger.info(f"  - Total procesados: {len(resultados)}")
        logger.info(f"  - Con teléfono: {sum(1 for t in resultados.values() if t and t != 'ERROR')}")
        
        return 0
        
    except Exception as e:
        logger.error(f"Error fatal: {e}", exc_info=True)
        return 1

if __name__ == "__main__":
    sys.exit(main())