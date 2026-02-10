"""Exporta siniestros Allianz desde PeritoLine (BD) a Excel y extrae teléfonos de ePAC.

FLUJO COMPLETO:
1. Extrae datos de BD → Excel con 4 columnas (Encargo, Fecha Sin., Causa, Aseguradora)
2. Se conecta a ePAC y extrae teléfonos
3. Actualiza Excel → 6 columnas (añade Teléfono y Estado)

Variables de entorno (o --env .env):
  DB_HOST, DB_NAME, DB_USER, DB_PASS, DB_SSL_CA, DB_SSL_CERT, DB_SSL_KEY
  EPAC_URL, EPAC_USERNAME, EPAC_PASSWORD (opcionales, se pueden obtener de BD)
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import mysql.connector
from mysql.connector import Error
from mysql.connector.constants import ClientFlag
from openpyxl import Workbook
from openpyxl.styles import Font
from dotenv import load_dotenv

# Cargar .env desde la raíz del proyecto
load_dotenv(dotenv_path=Path(__file__).resolve().parents[1] / ".env")

# Directorio de salida - ruta relativa al directorio actual
OUTPUT_DIR = Path("./data/peritoline/raw_allianz")

@dataclass
class DbConfig:
    host: str
    name: str
    user: str
    password: str


SQL_EXPORT = """
WITH
-- 1) Siniestros que "aparecen" en PL por tener algún encargo EN EL DÍA objetivo
siniestros_en_rango AS (
  SELECT
    e.id_siniestro,
    MAX(e.encargo_fh) AS encargo_fh_rango,
    MAX(CASE WHEN e.id_causa != 15 THEN e.id_causa ELSE NULL END) AS id_causa  -- Excluir "Asistencia a Juicio"
  FROM softline_encargos e
  WHERE e.id_despacho = 2
    AND e.encargo_fh >= %(day_start)s
    AND e.encargo_fh <  %(day_end)s
  GROUP BY e.id_siniestro
),

-- 2) Histórico completo de encargos SOLO para esos siniestros
enc_rank AS (
  SELECT
    e.id_siniestro,
    e.id_encargo,
    ROW_NUMBER() OVER (
      PARTITION BY e.id_siniestro
      ORDER BY e.encargo_fh, e.id_encargo
    ) AS rn,
    CASE
      WHEN (COALESCE(e.encargo_estado_ttr,0) > 0 OR e.encargo_tipo = 12)
      THEN 1 ELSE 0
    END AS is_ttr
  FROM softline_encargos e
  JOIN siniestros_en_rango sr ON sr.id_siniestro = e.id_siniestro
  WHERE e.id_despacho = 2
),

-- 3) Info del siniestro (nº encargos total + ttr_1/ttr_2 sobre histórico)
enc_info AS (
  SELECT
    id_siniestro,
    COUNT(*) AS n_encargos,
    MAX(CASE WHEN rn=1 THEN is_ttr ELSE 0 END) AS ttr_1,
    MAX(CASE WHEN rn=2 THEN is_ttr ELSE 0 END) AS ttr_2
  FROM enc_rank
  GROUP BY id_siniestro
)

SELECT
  c.codigo AS encargo,
  DATE(s.siniestro_fh) AS fecha_sin,
  COALESCE(mc.descripcion, '') AS causa,
  CASE 
    WHEN s.id_cia = 42 THEN 'Allianz'
    WHEN s.id_cia = 399 THEN 'AllianzBBVA'
    ELSE 'Desconocida'
  END AS aseguradora,
  COALESCE(i.nombre, '') AS asegurado,
  COALESCE(i.direccion, '') AS direccion,
  COALESCE(i.cp, '') AS codigo_postal,
  COALESCE(i.cia_municipio_str, '') AS municipio
FROM softline_siniestros s
JOIN siniestros_en_rango sr ON sr.id_siniestro = s.id_siniestro
JOIN softline_siniestros_codigos c
  ON c.id_codigo = s.id_cod_siniestro
 AND c.tipo_codigo = 'siniestro'
LEFT JOIN softline_maestra_causas mc
  ON mc.id_cod = sr.id_causa
LEFT JOIN softline_implicados i
  ON i.id_siniestro = s.id_siniestro
 AND i.tipo_implicado = 1
JOIN enc_info ei ON ei.id_siniestro = s.id_siniestro
WHERE s.id_despacho = 2
  AND s.id_cia IN (42,399)
  AND CHAR_LENGTH(c.codigo) >= 9

  -- sin contacto
  AND NOT EXISTS (
    SELECT 1
    FROM softline_encargos e
    JOIN softline_encargos_contactos ec
      ON ec.contactos_id_encargo = e.id_encargo
    WHERE e.id_siniestro = s.id_siniestro
      AND ec.contactos_fechahora IS NOT NULL
  )

  -- regla diagrama (sobre histórico)
  AND (
    (ei.n_encargos = 1 AND ei.ttr_1 = 0)
    OR
    (ei.n_encargos = 2 AND ei.ttr_2 = 0 AND ei.ttr_1 = 1)
  )
ORDER BY sr.encargo_fh_rango DESC, s.id_siniestro DESC;
"""


def load_env_file(env_path: Optional[str]) -> None:
    if not env_path:
        return
    p = Path(env_path)
    if not p.exists():
        raise FileNotFoundError(f"No existe el fichero --env: {p}")
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def get_db_config() -> DbConfig:
    host = os.environ.get("DB_HOST")
    name = os.environ.get("DB_NAME")
    user = os.environ.get("DB_USER")
    password = os.environ.get("DB_PASS")
    missing = [k for k, v in (('DB_HOST', host), ('DB_NAME', name), ('DB_USER', user), ('DB_PASS', password)) if not v]
    if missing:
        raise RuntimeError(f"Faltan variables de entorno para BD: {', '.join(missing)}")
    return DbConfig(host=host, name=name, user=user, password=password)


def day_bounds(target: date) -> tuple[datetime, datetime]:
    """Devuelve (inicio_inclusivo, fin_exclusivo) para un día concreto."""
    start = datetime(target.year, target.month, target.day, 0, 0, 0)
    end = start + timedelta(days=1)
    return start, end


def run_query(cfg: DbConfig, target_day: date) -> list[tuple[str, Optional[date], str, str, str, str, str, str]]:
    """Ejecuta la consulta SQL en la BD de PeritoLine usando mysql-connector-python con SSL.
    
    Args:
        cfg: Configuración de la base de datos.
        target_day: Día objetivo para filtrar encargos.
    
    Returns:
        Lista de tuplas (encargo, fecha_siniestro, causa, aseguradora, asegurado, direccion, codigo_postal, municipio).
    
    Raises:
        RuntimeError: Si hay error de conexión o ejecución.
    """
    day_start, day_end = day_bounds(target_day)
    
    # Parámetros SSL
    ssl_ca = os.environ.get("DB_SSL_CA")
    ssl_cert = os.environ.get("DB_SSL_CERT")
    ssl_key = os.environ.get("DB_SSL_KEY")
    
    try:
        # Configurar parámetros de conexión
        conn_params = {
            "host": cfg.host,
            "user": cfg.user,
            "password": cfg.password,
            "database": cfg.name,
            "connection_timeout": 10,
            "autocommit": True,
            "use_pure": True,
            "charset": "latin1",
        }
        
        if ssl_ca and ssl_cert and ssl_key:
            print(f"DEBUG: Configurando SSL con certificados")
            conn_params["client_flags"] = [ClientFlag.SSL]
            conn_params["ssl_ca"] = ssl_ca
            conn_params["ssl_cert"] = ssl_cert
            conn_params["ssl_key"] = ssl_key
            conn_params["ssl_disabled"] = False
            conn_params["ssl_verify_cert"] = True
            conn_params["ssl_verify_identity"] = False
        
        cnx = mysql.connector.connect(**conn_params)
        cur = cnx.cursor()
        cur.execute(SQL_EXPORT, {"day_start": day_start, "day_end": day_end})
        rows = cur.fetchall()
        cur.close()
        cnx.close()
        print(f"✓ Consulta exitosa: {len(rows)} filas")
        return [(str(r[0]) if r[0] is not None else "", r[1], str(r[2] or ""), str(r[3] or ""), str(r[4] or ""), str(r[5] or ""), str(r[6] or ""), str(r[7] or "")) for r in rows]
        
    except mysql.connector.errors.ProgrammingError as e:
        raise RuntimeError(f"Error de credenciales o permisos BD ({cfg.host}): {e}")
    except Error as e:
        raise RuntimeError(f"Error conectando con BD: {e}")


def write_excel(rows: list[tuple[str, Optional[date], str, str, str, str, str, str]], out_path: str) -> None:
    """Escribe el Excel con 8 columnas: Encargo, Fecha Sin., Causa, Aseguradora, Asegurado, Dirección, CP, Municipio"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Allianz"

    # Headers
    headers = ["Encargo", "Fecha Sin.", "Causa", "Aseguradora", "Asegurado", "Dirección", "CP", "Municipio"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data
    for i, (encargo, fecha, causa, aseguradora, asegurado, direccion, codigo_postal, municipio) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=encargo)
        
        c_fecha = ws.cell(row=i, column=2, value=fecha)
        if fecha is not None:
            c_fecha.number_format = "DD/MM/YYYY"
        
        ws.cell(row=i, column=3, value=causa)
        ws.cell(row=i, column=4, value=aseguradora)
        ws.cell(row=i, column=5, value=asegurado)
        ws.cell(row=i, column=6, value=direccion)
        ws.cell(row=i, column=7, value=codigo_postal)
        ws.cell(row=i, column=8, value=municipio)

    # Ajustar anchos
    ws.column_dimensions["A"].width = 16  # Encargo
    ws.column_dimensions["B"].width = 14  # Fecha Sin.
    ws.column_dimensions["C"].width = 30  # Causa
    ws.column_dimensions["D"].width = 16  # Aseguradora
    ws.column_dimensions["E"].width = 30  # Asegurado
    ws.column_dimensions["F"].width = 40  # Dirección
    ws.column_dimensions["G"].width = 10  # CP
    ws.column_dimensions["H"].width = 25  # Municipio

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--date", default=None, help="Día objetivo (YYYY-MM-DD). Por defecto: hoy")
    p.add_argument("--env", default=None, help="Ruta a fichero .env")
    p.add_argument("--skip-epac", action="store_true", help="Solo generar Excel, sin extraer teléfonos de ePAC")
    p.add_argument("--max", type=int, default=0, help="Procesar solo los primeros N siniestros en ePAC (0 = todos)")
    p.add_argument("--headless", action="store_true", help="Ejecutar ePAC en modo headless (sin navegador visible)")
    p.add_argument("--headed", action="store_true", help="Forzar modo con interfaz gráfica (requiere X server)")
    return p.parse_args()


def should_use_headless(args) -> bool:
    """Determina si debe usar modo headless automáticamente."""
    # Si el usuario especifica --headed, respetar esa decisión
    if args.headed:
        return False
    
    # Si el usuario especifica --headless, usarlo
    if args.headless:
        return True
    
    # Auto-detectar: si no hay DISPLAY, usar headless
    if not os.environ.get('DISPLAY'):
        print("   ℹ️  No se detectó display gráfico, usando modo headless automáticamente")
        return True
    
    return False


def main() -> None:
    args = parse_args()
    load_env_file(args.env)

    target_day = date.today() if not args.date else date.fromisoformat(args.date)
    
    # Crear directorio de salida
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Solo crear allianz_latest.xlsx (sin timestamp)
    latest_path = OUTPUT_DIR / "allianz_latest.xlsx"
    
    print(f"📊 PASO 1/2: Extrayendo datos de la base de datos...")
    print(f"   Fecha objetivo: {target_day}")
    
    cfg = get_db_config()
    rows = run_query(cfg, target_day)
    
    if not rows:
        print(f"⚠️  No se encontraron siniestros para la fecha {target_day}")
        print(f"   Verifica que haya datos en la BD para esa fecha")
        return
    
    write_excel(rows, str(latest_path))
    
    print(f"✅ Excel generado: {latest_path}")
    print(f"   → {len(rows)} siniestros exportados")
    print(f"   → 8 columnas: Encargo, Fecha Sin., Causa, Aseguradora, Asegurado, Dirección, CP, Municipio")
    
    # PASO 2: Extraer teléfonos de ePAC (si no se especifica --skip-epac)
    if not args.skip_epac:
        print(f"\n📞 PASO 2/2: Extrayendo teléfonos de ePAC...")
        
        try:
            # Construir comando para ejecutar el script de ePAC
            script_dir = Path(__file__).resolve().parent
            epac_script = script_dir / "extraer_teléfonos_epac_updated.py"
            
            if not epac_script.exists():
                epac_script = script_dir / "extraer_teléfonos_epac.py"
            
            if not epac_script.exists():
                print(f"⚠️  No se encontró el script de ePAC")
                print(f"   Excel generado en: {excel_path}")
                print(f"   Ejecuta manualmente: python extraer_teléfonos_epac.py --excel {excel_path}")
                return
            
            cmd = [sys.executable, str(epac_script), "--excel", str(latest_path)]
            
            if args.max > 0:
                cmd.extend(["--max", str(args.max)])
            
            # Auto-detectar o usar argumento de headless
            if should_use_headless(args):
                cmd.append("--headless")
            
            # Ejecutar script de ePAC (stdout en tiempo real, capturar stderr)
            result = subprocess.run(cmd, stderr=subprocess.PIPE, text=True)

            if result.returncode == 0:
                print(f"✅ Teléfonos extraídos correctamente")
                print(f"   → Excel actualizado con columnas: Teléfono y Estado")
                print(f"\n📁 Archivo final: {latest_path}")
            else:
                print(f"❌ Error al extraer teléfonos (código {result.returncode}):")
                if result.stderr:
                    print(result.stderr)
                print(f"\n📁 Excel (sin teléfonos): {latest_path}")
                
        except Exception as e:
            print(f"❌ Error al ejecutar extracción de ePAC: {e}")
            print(f"📁 Excel (sin teléfonos): {latest_path}")
    else:
        print(f"\n⏭️  Extracción de teléfonos omitida (--skip-epac)")
        print(f"📁 Excel generado: {latest_path}")


if __name__ == "__main__":
    main()