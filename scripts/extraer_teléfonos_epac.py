"""
Extrae teléfonos desde ePAC para siniestros Allianz.

Modos:
- --refresh: borra excels existentes (raw_allianz) y descarga de nuevo desde Peritoline
- --excel: usa un Excel existente (sin descargar)
- --headless: ejecuta con navegador oculto (sin UI). Si no, con navegador visible.

Flujo:
1) Generar/obtener Excel desde BD (export_allianz_from_db) o usar uno existente
   El Excel contiene: Encargo, Fecha Sin., Causa, Aseguradora
2) Obtener credenciales ePAC (preferencia: BD/CLI/env)
3) Login ePAC -> ir a Peritaciones Diversos
4) Por cada siniestro:
   - asegurar pantalla de búsqueda (#claimNumber visible)
   - buscar, seleccionar
   - menú lateral: HSC y FP -> Ficha peritación
   - extraer teléfono (obs -> descripción -> telef-1 -> telef-2)
   - volver a Peritaciones Diversos
5) Actualizar el Excel con: Teléfono y Estado
   Columnas finales: Encargo, Fecha Sin., Causa, Aseguradora, Teléfono, Estado
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from playwright.sync_api import Page  # pragma: no cover

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config import AppConfig, load_config
from export_allianz_from_db import get_db_config as _get_db_config, run_query as _run_query, write_excel as _write_excel


from epac.pages.epac_ficha_peritacion_page import EpacFichaPeritacionPage
from epac.pages.login_page import LoginPage
from epac.pages.menu_lateral_page import MenuLateralPage
from epac.pages.navigation_page import NavigationPage
from epac.pages.num_siniestro_page import NumeroSiniestroPage
from utils.human import human_delay
from utils.logging_utils import get_logger, setup_logging

# Importar mysql.connector para obtener credenciales de BD (sin Playwright)
try:
    import mysql.connector  # type: ignore
except ImportError:
    mysql = None  # noqa: F401
    mysql_import_error = ImportError("mysql.connector no disponible")
else:
    mysql_import_error = None

# Importar openpyxl para manipular Excel
import openpyxl
from openpyxl.styles import Font


RAW_DIR = Path("data/peritoline/raw_allianz")


# -----------------------------------------------------------------------------
# Utilidades Excel
# -----------------------------------------------------------------------------

def borrar_excels_raw_allianz() -> int:
    """Borra los excels de salida en el directorio RAW.

    Returns:
        Numero de archivos eliminados.

    Notes:
        Documentación pensada para MkDocs.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    n = 0
    for p in RAW_DIR.glob("*.xlsx"):
        try:
            p.unlink()
            n += 1
        except Exception:
            pass
    for p in RAW_DIR.glob("*.xls"):
        try:
            p.unlink()
            n += 1
        except Exception:
            pass
    return n


def pick_latest_excel(raw_dir: Path) -> Optional[Path]:
    """Devuelve el Excel más reciente dentro de un directorio.

    Args:
        raw_dir: Directorio donde buscar archivos.

    Returns:
        Ruta del Excel más reciente o None si no hay.

    Notes:
        Documentación pensada para MkDocs.
    """
    files = [p for p in raw_dir.glob("*.xlsx") if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def normalizar_siniestro(s: str) -> str:
    """Normaliza un siniestro dejando solo dígitos.

    Args:
        s: Valor a normalizar.

    Returns:
        Cadena con solo dígitos.

    Notes:
        Documentación pensada para MkDocs.
    """
    return re.sub(r"\D", "", (s or "").strip())


def filtrar_siniestros_validos(lista: list[str], min_len: int = 9) -> list[str]:
    """Filtra siniestros válidos por longitud y unicidad.

    Args:
        lista: Lista de siniestros en bruto.
        min_len: Longitud mínima requerida.

    Returns:
        Lista de siniestros normalizados y válidos.

    Notes:
        Documentación pensada para MkDocs.
    """
    out = []
    seen = set()
    for s in lista:
        x = normalizar_siniestro(s)
        if len(x) < min_len:
            continue
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


# -----------------------------------------------------------------------------
# Descarga + credenciales (Peritoline)
# -----------------------------------------------------------------------------

def _cerrar_popup_avisos_si_existe(page: Page, config: AppConfig) -> None:
    """Cierra el popup de avisos si aparece en Peritoline.

    Args:
        page: Pagina activa de Playwright.
        config: Configuración de la aplicación.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    boton_ok = page.locator("button.swal2-confirm.swal2-styled:has-text('OK')")
    try:
        if boton_ok.is_visible():
            human_delay(config, "Cerrando popup de avisos")
            boton_ok.click()
    except Exception:
        return


def descargar_excel_allianz(config: AppConfig) -> Path:
    """Descarga el Excel de Allianz desde Peritoline.

    Args:
        config: Configuración de la aplicación.

    Returns:
        Ruta del Excel descargado.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(tarea="descarga_excel")

    # Lazy imports: solo si descargamos desde Peritoline
    from browser import launch_browser
    from peritoline.login_page import PeritolineLoginPage
    from peritoline.downloader import download_report_in_session
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = RAW_DIR / f"allianz_{timestamp}.xlsx"
    latest_path = RAW_DIR / "allianz_latest.xlsx"

    with launch_browser(config) as (_, page):
        human_delay(config, "Login Peritoline")
        login_page = PeritolineLoginPage(page)
        login_page.open(config.peritoline_login_url)  # <- URL string, no config
        login_page.login(config.peritoline_username, config.peritoline_password)
        _cerrar_popup_avisos_si_existe(page, config)

        download_report_in_session(
            page,
            config,
            output_path,
            "ALLIANZ",
            click_view_all_button=True,
            click_view_all_list=True,
        )

    shutil.copy2(output_path, latest_path)
    logger.info(f"Excel descargado en: {output_path}")
    return output_path


def obtener_credenciales_epac(config: AppConfig) -> dict:
    """Obtiene credenciales de ePAC desde la base de datos o entorno.

    Args:
        config: Configuración de la aplicación.

    Returns:
        Diccionario con url, username y password.

    Notes:
        Documentación pensada para MkDocs.
        La BD usa softline_aseguradoras_claves_web. En caso de fallo se
        intenta con config y variables EPAC_*.
    """
    logger = get_logger(tarea="obtener_credenciales_epac")
    logger.info("Obteniendo credenciales de ePAC desde la base de datos")

    # 1) Intento BD
    if mysql_import_error is None:
        db_host = getattr(config, "db_host", None) or os.getenv("DB_HOST")
        db_port = int(getattr(config, "db_port", None) or os.getenv("DB_PORT", "3306"))
        db_user = getattr(config, "db_user", None) or os.getenv("DB_USER")
        db_password = getattr(config, "db_password", None) or os.getenv("DB_PASSWORD") or os.getenv("DB_PASS")
        db_name = getattr(config, "db_name", None) or os.getenv("DB_NAME") or "criteria_peritoline"

        if db_host and db_user and db_password and db_name:
            cnx = None
            try:
                # Configurar parámetros SSL si existen
                conn_params = {
                    "host": db_host,
                    "port": db_port,
                    "user": db_user,
                    "password": db_password,
                    "database": db_name,
                    "connection_timeout": 10,
                    "autocommit": True,
                    "use_pure": True,
                    "charset": "latin1",
                }
                
                # Parámetros SSL
                if os.environ.get("DB_SSL_CA") or os.environ.get("DB_SSL_CERT") or os.environ.get("DB_SSL_KEY"):
                    conn_params["ssl_disabled"] = False
                    ssl_config = {}
                    if os.environ.get("DB_SSL_CA"):
                        ssl_config["ca"] = os.environ.get("DB_SSL_CA")
                    if os.environ.get("DB_SSL_CERT"):
                        ssl_config["cert"] = os.environ.get("DB_SSL_CERT")
                    if os.environ.get("DB_SSL_KEY"):
                        ssl_config["key"] = os.environ.get("DB_SSL_KEY")
                    conn_params["ssl_verify_cert"] = False
                    conn_params["ssl_verify_identity"] = False
                    conn_params["ssl_ca"] = ssl_config.get("ca")
                    conn_params["ssl_cert"] = ssl_config.get("cert")
                    conn_params["ssl_key"] = ssl_config.get("key")
                
                cnx = mysql.connector.connect(**conn_params)  # type: ignore[name-defined]
                query = (
                    "SELECT url, user AS usuario, pass AS password "
                    "FROM softline_aseguradoras_claves_web "
                    "WHERE id_cia IN (42, 399) "
                    "LIMIT 1"
                )
                cur = cnx.cursor(dictionary=True)
                cur.execute(query)
                row = cur.fetchone()
                cur.close()
                cnx.close()

                if row:
                    logger.info("Credenciales obtenidas desde BD")
                    return {
                        "url": row["url"],
                        "username": row["usuario"],
                        "password": row["password"],
                    }
            except Exception as e:
                logger.warning(f"No pude obtener credenciales desde BD: {e}")

    # 2) Fallback a config
    epac_url = getattr(config, "epac_url", None) or os.getenv("EPAC_URL")
    epac_username = getattr(config, "epac_username", None) or os.getenv("EPAC_USERNAME")
    epac_password = getattr(config, "epac_password", None) or os.getenv("EPAC_PASSWORD")

    if epac_url and epac_username and epac_password:
        logger.info("Usando credenciales de config/env")
        return {"url": epac_url, "username": epac_username, "password": epac_password}

    raise RuntimeError(
        "No se pudieron obtener credenciales de ePAC. "
        "Verifica BD o variables de entorno: EPAC_URL, EPAC_USERNAME, EPAC_PASSWORD"
    )


# -----------------------------------------------------------------------------
# ePAC: navegación / extracción
# -----------------------------------------------------------------------------

EPAC_PRIVATE_APP_URL = "https://www.e-pacallianz.com/ngx-epac-professional/private/"


def login_epac(page: Page, url: str, usuario: str, password: str) -> None:
    """Inicia sesion en ePAC usando el formulario de login.

    Args:
        page: Pagina activa de Playwright.
        url: URL base de ePAC.
        usuario: Usuario de acceso.
        password: Clave de acceso.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(tarea="login_epac")
    logger.info("Login en ePAC")
    login_page = LoginPage(page)
    login_page.open(url)
    login_page.login(usuario, password)
    page.wait_for_timeout(1000)


def navegar_a_peritaciones_diversos(page: Page, config: AppConfig) -> None:
    """Navega al formulario de Peritaciones Diversos.

    Args:
        page: Pagina activa de Playwright.
        config: Configuración de la aplicación.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(tarea="navegar_peritaciones_diversos")
    logger.info("Navegando a Peritaciones Diversos")
    navigation = NavigationPage(page, config)
    navigation.goto_informe_pericial_diversos_sea()
    siniestro_page = NumeroSiniestroPage(page)
    siniestro_page.wait_until_ready()


def asegurar_pantalla_busqueda(
    page: Page,
    config: AppConfig,
    reintentos: int = 3,
) -> None:
    """Asegura que #claimNumber esté visible. Si no, navega.

    Args:
        page: Pagina activa de Playwright.
        config: Configuración de la aplicación.
        reintentos: Numero de reintentos.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    for intento in range(1, reintentos + 1):
        try:
            siniestro_page = NumeroSiniestroPage(page)
            siniestro_page.wait_until_ready()
            return
        except Exception:
            if intento < reintentos:
                navegar_a_peritaciones_diversos(page, config)
            else:
                raise RuntimeError("No pude asegurar pantalla de búsqueda")


def abrir_ficha_peritacion_menu_lateral(
    page: Page,
    siniestro: str,
    config: AppConfig,
    reintentos: int = 3,
) -> None:
    """Abre la ficha de peritacion desde el menu lateral.

    Args:
        page: Pagina activa de Playwright.
        siniestro: Codigo del siniestro.
        config: Configuración de la aplicación.
        reintentos: Numero de reintentos.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(siniestro=siniestro, tarea="abrir_ficha_menu")
    for intento in range(1, reintentos + 1):
        try:
            menu_page = MenuLateralPage(page, siniestro=siniestro)
            menu_page.abrir_ficha_peritacion()
            return
        except Exception as e:
            logger.warning(f"Intento {intento}/{reintentos} fallo: {e}")
            if intento < reintentos:
                asegurar_pantalla_busqueda(page, config, reintentos=1)
                page.wait_for_timeout(1000)
            else:
                raise


def volver_a_busqueda_desde_ficha(page: Page, config: AppConfig) -> None:
    """Vuelve a la pantalla de búsqueda desde la ficha si es posible.

    Args:
        page: Pagina activa de Playwright.
        config: Configuración de la aplicación.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(tarea="volver_busqueda")
    try:
        volver_btn = page.locator("button:has-text('Volver a búsqueda')").first
        volver_btn.wait_for(state="visible", timeout=5000)
        volver_btn.click()
        page.wait_for_timeout(300)
        logger.info("Vuelto a búsqueda")
    except Exception as e:
        logger.warning(f"No pude volver con botón: {e}. Navegando manualmente.")
        navegar_a_peritaciones_diversos(page, config)


# -----------------------------------------------------------------------------
# Extracción teléfonos
# -----------------------------------------------------------------------------

PHONE_CANDIDATE_RE = re.compile(
    r"""
    (?:
        (?:\+|00)\s*\d{1,3}[\s\-\.]?  # prefijo internacional
    )?
    \(?\d{1,4}\)?                    # área
    [\s\-\.\,]*                      # separadores
    \d{1,4}                          # grupo
    [\s\-\.\,]*
    \d{1,4}
    (?:[\s\-\.\,]*\d{1,4})?
    """,
    re.VERBOSE,
)


def normalizar_telefono(s: str) -> Optional[str]:
    """Normaliza teléfonos en formato español o internacional.

    Args:
        s: Valor de teléfono a normalizar.

    Returns:
        Teléfono normalizado o None si no es válido.

    Notes:
        Documentación pensada para MkDocs.
    """
    if not s:
        return None

    if s.startswith("+"):
        digits = "+" + re.sub(r"\D", "", s[1:])
    else:
        digits = re.sub(r"\D", "", s)

    if digits in ("", "+"):
        return None

    # España con +34
    if digits.startswith("+34"):
        rest = digits[3:]
        if len(rest) == 9 and rest[0] in "6789":
            return rest

    # España con 34 sin +
    if not digits.startswith("+") and digits.startswith("34") and len(digits) == 11 and digits[2] in "6789":
        return digits[-9:]

    # 0 delante
    if not digits.startswith("+") and len(digits) == 10 and digits.startswith("0") and digits[1] in "6789":
        return digits[1:]

    # España directa
    if not digits.startswith("+") and len(digits) == 9 and digits[0] in "6789":
        return digits

    # Internacional
    if digits.startswith("+"):
        if 10 <= len(digits) - 1 <= 15:
            return digits
        return None

    if 10 <= len(digits) <= 15:
        return digits

    return None


def extraer_telefono_de_texto(texto: str) -> Optional[str]:
    """Extrae el primer teléfono válido detectado en un texto.

    Args:
        texto: Texto origen.

    Returns:
        Teléfono normalizado o None si no hay coincidencias.

    Notes:
        Documentación pensada para MkDocs.
    """
    if not texto:
        return None
    for m in PHONE_CANDIDATE_RE.finditer(texto):
        tel = normalizar_telefono(m.group(0))
        if tel:
            return tel
    return None


# -----------------------------------------------------------------------------
# Procesado por siniestro
# -----------------------------------------------------------------------------

def procesar_siniestro(page: Page, numero_siniestro: str, config: AppConfig) -> dict:
    """Procesa un siniestro y devuelve el resultado de telefono.

    Args:
        page: Pagina activa de Playwright.
        numero_siniestro: Codigo del siniestro a procesar.
        config: Configuración de la aplicación.

    Returns:
        Diccionario con siniestro, telefono y estado.

    Notes:
        Documentación pensada para MkDocs.
    """
    logger = get_logger(siniestro=numero_siniestro, tarea="procesar_siniestro")
    logger.info("Procesando siniestro")

    try:
        # 0) asegurar pantalla búsqueda
        asegurar_pantalla_busqueda(page, config, reintentos=3)

        # 1) Buscar siniestro
        siniestro_page = NumeroSiniestroPage(page)
        siniestro_page.wait_until_ready()
        siniestro_page.fill_siniestro_number(numero_siniestro)
        siniestro_page.submit_codigo()

        # 2) Seleccionar resultado
        siniestro_page.seleccionar_resultado_por_codigo(numero_siniestro)

        # 3) Ir a ficha peritación por menú lateral
        frame = page.frame_locator("iframe[name='appArea']")
        frame.locator("body").wait_for(state="attached", timeout=15000)
        abrir_ficha_peritacion_menu_lateral(page, numero_siniestro, config, reintentos=3)

        # Espera a que la ficha termine de renderizar dentro del iframe
        frame.locator("body").wait_for(state="attached", timeout=15000)

        # 4) Extraer teléfono SIN cambiar de pantalla
        ficha = EpacFichaPeritacionPage(page)
        telefono = ficha.extraer_telefono()
        logger.info(f"Teléfono: {telefono or 'NO ENCONTRADO'}")

        return {
            "siniestro": numero_siniestro,
            "telefono": telefono,
            "estado": "OK" if telefono else "NO ENCONTRADO",
        }

    except Exception as e:
        logger.error(f"Error: {e}")
        return {
            "siniestro": numero_siniestro,
            "telefono": None,
            "estado": "ERROR",
            "error": str(e),
        }


def actualizar_excel_con_telefonos(excel_path: Path, resultados: list[dict]) -> None:
    """Actualiza el Excel con columnas Teléfono y Estado.

    Args:
        excel_path: Ruta del Excel a actualizar.
        resultados: Lista de resultados por siniestro.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Leer headers
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    
    # Añadir nuevas columnas si no existen
    if "Teléfono" not in headers:
        col_telefono = len(headers) + 1
        ws.cell(row=1, column=col_telefono, value="Teléfono")
        ws.cell(row=1, column=col_telefono).font = Font(bold=True)
    else:
        col_telefono = headers.index("Teléfono") + 1
    
    if "Estado" not in headers:
        col_estado = len(headers) + 2 if "Teléfono" not in headers else len(headers) + 1
        ws.cell(row=1, column=col_estado, value="Estado")
        ws.cell(row=1, column=col_estado).font = Font(bold=True)
    else:
        col_estado = headers.index("Estado") + 1
    
    # Mapear siniestros a resultados
    resultados_map = {normalizar_siniestro(r["siniestro"]): r for r in resultados}
    
    # Actualizar filas
    encargo_col = headers.index("Encargo") + 1 if "Encargo" in headers else 1
    
    for row_idx in range(2, ws.max_row + 1):
        encargo_cell = ws.cell(row=row_idx, column=encargo_col)
        encargo = normalizar_siniestro(str(encargo_cell.value or ""))
        
        if encargo in resultados_map:
            resultado = resultados_map[encargo]
            ws.cell(row=row_idx, column=col_telefono, value=resultado.get("telefono") or "")
            ws.cell(row=row_idx, column=col_estado, value=resultado.get("estado") or "")
    
    # Ajustar anchos
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_telefono)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_estado)].width = 15
    
    wb.save(excel_path)
    print(f"Excel actualizado: {excel_path}")


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------

def exportar_excel_desde_bd() -> Path:
    """Genera el Excel de Allianz desde la BD con export_allianz_from_db.

    Returns:
        Ruta del archivo Excel generado.

    Notes:
        Documentación pensada para MkDocs.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    out = RAW_DIR / "allianz_report_latest.xlsx"
    # export_allianz_from_db lee DB_* desde .env / entorno (ya cargado por load_config)
    cfg = _get_db_config()
    target_day = datetime.now().date()
    rows = _run_query(cfg, target_day)
    _write_excel(rows, out)
    return out


def main() -> None:
    """Ejecuta el flujo de extracción de teléfonos ePAC.

    Returns:
        None.

    Notes:
        Documentación pensada para MkDocs.
    """
    setup_logging()
    logger = get_logger(tarea="extraer_telefonos_epac")

    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", nargs="?", const="__AUTO__", default=None,
                        help="Ruta a un Excel existente. Si se usa sin valor, se usará el último generado en RAW_DIR.")
    parser.add_argument("--refresh", action="store_true",
                        help="Borra excels existentes en data/peritoline/raw_allianz y descarga uno nuevo.")
    parser.add_argument("--headless", action="store_true",
                        help="Ejecutar en modo headless (sin abrir navegador). Si no se indica, se ve el navegador.")
    parser.add_argument("--min-siniestro-len", type=int, default=9,
                        help="Longitud mínima de siniestro numérico para procesar (por defecto 9).")
    parser.add_argument("--max", type=int, default=0,
                        help="Procesa solo los primeros N siniestros (0 = todos).")
    args = parser.parse_args()

    config = load_config()
    config.headless = args.headless

    # 1) Excel (preferencia: BD). Si no se indica --excel, generamos/recogemos uno en RAW_DIR.
    excel_in: Optional[Path] = None
    if args.refresh:
        borrados = borrar_excels_raw_allianz()
        logger.info(f"Refresh activado: excels borrados={borrados}")
        excel_in = exportar_excel_desde_bd()
    elif args.excel is not None:
        excel_in = pick_latest_excel(RAW_DIR) if args.excel == "__AUTO__" else Path(args.excel)
        if not excel_in.exists():
            raise SystemExit(f"No existe el excel indicado: {excel_in}")
    else:
        excel_in = pick_latest_excel(RAW_DIR)
        if not excel_in:
            excel_in = exportar_excel_desde_bd()

    print(f"Excel (BD) guardado en: {excel_in}")

    # 2) Leer siniestros/encargos (columna "Encargo")
    wb = openpyxl.load_workbook(excel_in)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if "Encargo" not in headers:
        raise SystemExit("No encuentro columna 'Encargo' en el excel.")
    idx = headers.index("Encargo")
    siniestros: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        siniestros.append(str(row[idx] or "").strip())

    siniestros = filtrar_siniestros_validos(siniestros, min_len=args.min_siniestro_len)
    if args.max and args.max > 0:
        siniestros = siniestros[:args.max]

    print(f"  - Siniestros válidos: {len(siniestros)}")

    if not siniestros:
        print("No hay siniestros válidos tras filtrar.")
        return

    # 4) Credenciales ePAC
    cred = obtener_credenciales_epac(config)
    url_epac = cred.get("url") or "https://www.e-pacallianz.com/ngx-epac-professional/"

    # 5) ePAC extracción
    resultados: list[dict] = []

    from browser import launch_browser  # Lazy import: solo al entrar en ePAC
    with launch_browser(config) as (_, page):
        login_epac(page, url_epac, cred["username"], cred["password"])
        navegar_a_peritaciones_diversos(page, config)

        total = len(siniestros)
        for i, s in enumerate(siniestros, start=1):
            print(f"[{i}/{total}] {s}")
            resultados.append(procesar_siniestro(page, s, config))

            # volver siempre a Peritaciones Diversos de forma robusta
            try:
                volver_a_busqueda_desde_ficha(page, config)
            except Exception:
                try:
                    page.goto(EPAC_PRIVATE_APP_URL, wait_until="domcontentloaded")
                    page.wait_for_timeout(800)
                    navegar_a_peritaciones_diversos(page, config)
                except Exception:
                    pass

    # 6) Actualizar Excel con teléfonos
    actualizar_excel_con_telefonos(excel_in, resultados)
    print("Proceso completado. Excel actualizado con teléfonos.")


if __name__ == "__main__":
    main()
